import openpyxl
import re

wb = openpyxl.load_workbook('C:\\Users\\srikar.kadavakuti\\Desktop\\June17-Project\\alternate.xlsx')
wb2 = openpyxl.load_workbook('C:\\Users\\srikar.kadavakuti\\Desktop\\June17-Project\\outsource.xlsx')
#print(wb.sheetnames)

# This is Alternate list section -> Load the part numbers are store in a list named alt_list
alt_list = []
sheet = wb['Sheet1']

for rowNum in range(7, sheet.max_row+1):
    partNum = sheet.cell(row=rowNum, column = 1).value
    if partNum != None:
        alt_list.append(partNum)

alt_part_list = []
for i in alt_list:
    i = str(i)
    if re.match(r'\d{8}', i):
        alt_part_list.append(i)
#print(alt_part_list)

# This is Outsource list section

out_list = []
out_sheet = wb2['7845']
for rowNum in range(11, out_sheet.max_row+1):
    pNum = out_sheet.cell(row=rowNum, column=10).value
    if pNum != None:
        out_list.append(pNum)

out_part_list = []
for i in out_list:
    i = str(i)
    if re.match(r'\d{8}', i):
        out_part_list.append(i)
#print(out_part_list)

uncommon_list = []
common_list = []
for i in alt_part_list:
    if i in out_part_list:
        common_list.append(i)
    else:
        uncommon_list.append(i)
#print("Uncommon List: ", uncommon_list)

wb.create_sheet(title='Uncommon Parts')
sheet = wb['Uncommon Parts']
for j in uncommon_list:
    sheet['A1'+ str(1)] = j
wb.save('C:\\Users\\srikar.kadavakuti\\Desktop\\June17-Project\\alternate.xlsx')

